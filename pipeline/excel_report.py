"""Excel report automation for CryptoFlow using openpyxl."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from pipeline.api_ingest import DB_PATH, get_latest_market_data, run_market_ingest
from pipeline.cohort_analysis import analyze_cohorts
from pipeline.funnel_analysis import DATA_PATH, analyze_funnel, load_funnel_events


OUTPUT_PATH = PROJECT_ROOT / "outputs" / "CryptoFlow_Report.xlsx"


def style_header(row_cells: tuple) -> None:
    """Apply a consistent bold header style to an openpyxl row."""
    for cell in row_cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A73E8")
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(worksheet) -> None:
    """Set worksheet column widths based on visible cell contents."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)


def add_funnel_sheet(workbook: Workbook) -> None:
    """Create the Funnel Analysis worksheet with formatting and formulas."""
    funnel, metrics = analyze_funnel(print_summary=False, return_metrics=True)
    worksheet = workbook.active
    worksheet.title = "Funnel Analysis"

    headers = ["Step", "Users Reached", "% of Signups", "Drop Rate", "Status"]
    worksheet.append(headers)
    style_header(tuple(worksheet[1]))

    for _, row in funnel.iterrows():
        status = "Anomaly" if bool(row["is_anomaly"]) else "Healthy"
        worksheet.append(
            [
                row["step_name"],
                int(row["users_reached"]),
                float(row["pct_of_signups"]),
                float(row["drop_rate"]),
                status,
            ]
        )

    worksheet.freeze_panes = "A2"
    red_fill = PatternFill("solid", fgColor="FFCCCC")
    amber_fill = PatternFill("solid", fgColor="FFF3CC")
    green_fill = PatternFill("solid", fgColor="CCFFCC")
    worksheet.conditional_formatting.add("D2:D9", CellIsRule(operator="greaterThan", formula=["15"], fill=red_fill))
    worksheet.conditional_formatting.add("D2:D9", CellIsRule(operator="between", formula=["8", "15"], fill=amber_fill))
    worksheet.conditional_formatting.add("D2:D9", CellIsRule(operator="lessThan", formula=["8"], fill=green_fill))

    worksheet["G2"] = "Overall KYC Completion"
    worksheet["H2"] = f"{metrics['kyc_completion_rate']:.1f}%"
    worksheet["G3"] = "Highest Drop Step"
    worksheet["H3"] = "=INDEX(A2:A9,MATCH(MAX(D2:D9),D2:D9,0))"
    worksheet["G4"] = "INDEX/MATCH Formula"
    worksheet["H4"] = "Finds the step with the highest drop rate"
    worksheet["G6"] = "Device Pivot"
    worksheet["G6"].font = Font(bold=True)
    worksheet.append([])

    events = load_funnel_events(DATA_PATH)
    device_counts = events.loc[events["step_name"] == "email_signup"].groupby("device_type")["user_id"].nunique()
    start_row = 8
    worksheet.cell(start_row, 7, "Device")
    worksheet.cell(start_row, 8, "Signups")
    worksheet.cell(start_row, 9, "Share")
    style_header((worksheet.cell(start_row, 7), worksheet.cell(start_row, 8), worksheet.cell(start_row, 9)))
    total = float(device_counts.sum() or 1)
    for index, (device, count) in enumerate(device_counts.items(), start=start_row + 1):
        worksheet.cell(index, 7, device)
        worksheet.cell(index, 8, int(count))
        worksheet.cell(index, 9, round(float(count) / total * 100, 1))

    autosize_columns(worksheet)


def retention_fill(value: float) -> PatternFill:
    """Return the heatmap fill color for a retention percentage."""
    if value > 60:
        return PatternFill("solid", fgColor="1B5E20")
    if value >= 40:
        return PatternFill("solid", fgColor="A5D6A7")
    if value >= 25:
        return PatternFill("solid", fgColor="FFF3CC")
    return PatternFill("solid", fgColor="FFCCCC")


def add_cohort_sheet(workbook: Workbook) -> None:
    """Create the Cohort Retention worksheet with a color heatmap."""
    cohort_df, _ = analyze_cohorts(print_summary=False)
    worksheet = workbook.create_sheet("Cohort Retention")
    worksheet.append(["Cohort Week"] + [f"W{week}" for week in range(8)])
    style_header(tuple(worksheet[1]))

    for _, row in cohort_df.iterrows():
        worksheet.append([row["cohort_week"]] + [float(row[f"W{week}"]) for week in range(8)])

    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=9):
        row[0].font = Font(bold=True)
        for cell in row[1:]:
            cell.fill = retention_fill(float(cell.value))
            cell.alignment = Alignment(horizontal="center")
            if float(cell.value) > 60:
                cell.font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "B2"
    autosize_columns(worksheet)


def latest_market_rows() -> list[dict]:
    """Load latest market data from SQLite, seeding the database if necessary."""
    if not DB_PATH.exists():
        run_market_ingest()
    rows = get_latest_market_data(DB_PATH)
    if not rows:
        run_market_ingest()
        rows = get_latest_market_data(DB_PATH)
    return rows


def add_market_sheet(workbook: Workbook) -> None:
    """Create the Market Data worksheet with conditional formatting and a bar chart."""
    rows = latest_market_rows()
    worksheet = workbook.create_sheet("Market Data")
    worksheet.append(["Name", "Symbol", "Price (INR)", "Market Cap", "24h Change%", "24h Volume"])
    style_header(tuple(worksheet[1]))

    for row in rows:
        worksheet.append(
            [
                row["name"],
                row["symbol"],
                float(row["current_price"]),
                float(row["market_cap"]),
                float(row["price_change_24h"]),
                float(row["total_volume"]),
            ]
        )

    green_font = Font(color="008000")
    red_font = Font(color="CC0000")
    worksheet.conditional_formatting.add("E2:E11", CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=green_font))
    worksheet.conditional_formatting.add("E2:E11", CellIsRule(operator="lessThan", formula=["0"], font=red_font))

    chart = BarChart()
    chart.title = "24h Trading Volume"
    chart.y_axis.title = "Volume (INR)"
    chart.x_axis.title = "Coin"
    data = Reference(worksheet, min_col=6, min_row=1, max_row=min(11, len(rows) + 1))
    categories = Reference(worksheet, min_col=2, min_row=2, max_row=min(11, len(rows) + 1))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    worksheet.add_chart(chart, "H2")
    autosize_columns(worksheet)


def create_excel_report(output_path: Path = OUTPUT_PATH) -> Path:
    """Build and save the full CryptoFlow Excel workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    add_funnel_sheet(workbook)
    add_cohort_sheet(workbook)
    add_market_sheet(workbook)
    workbook.save(output_path)
    print("Excel report generated: outputs/CryptoFlow_Report.xlsx")
    return output_path


def main() -> None:
    """CLI entry point for Excel report generation."""
    create_excel_report()


if __name__ == "__main__":
    main()
